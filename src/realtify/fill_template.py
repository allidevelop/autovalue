from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from pydantic import BaseModel, Field
from rich.console import Console

from realtify.excel_sidecar import build_calculation_sidecar_payload, write_excel_sidecar
from realtify.excel_tools import ExcelApp, excel_com_available, excel_path, libreoffice_available, libreoffice_path
from realtify.models import Comparable
from realtify.paths import PROJECT_ROOT, RESOURCE_ROOT, ensure_output_dir


class TemplateProfile(BaseModel):
    profile: str
    sheet_name: str
    fill_range: str
    comparables_columns: list[str]
    target_column: str | None = None
    protected_cells: list[str] = Field(default_factory=list)
    field_mapping: dict[str, int]
    adjustment_mapping: dict[str, int] = Field(default_factory=dict)
    hyperlink_rows: list[int] = Field(default_factory=list)


class FillError(RuntimeError):
    pass


@dataclass(frozen=True)
class FillResult:
    output_path: Path
    filled_count: int
    warnings: list[str]
    engine: str = ""
    metadata_path: Path | None = None


def load_template_profile(path: Path) -> TemplateProfile:
    with path.open("r", encoding="utf-8") as f:
        return TemplateProfile.model_validate(yaml.safe_load(f))


def load_candidates(path: Path) -> list[Comparable]:
    with path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    raw_candidates = payload.get("candidates")
    if not isinstance(raw_candidates, list):
        raise FillError(f"{path} does not contain a candidates list")
    return [Comparable.model_validate(item) for item in raw_candidates]


def fill_excel_template(
    *,
    template_path: Path,
    profile: TemplateProfile,
    candidates: list[Comparable],
    output_path: Path,
    target: dict[str, Any] | None = None,
    required_count: int | None = None,
    allow_less: bool = False,
    allow_incomplete: bool = False,
    visible: bool = False,
) -> FillResult:
    required = required_count or len(profile.comparables_columns)
    warnings = _validate_candidates(
        candidates,
        required_count=required,
        allow_less=allow_less,
        allow_incomplete=allow_incomplete,
    )
    selected = candidates[: len(profile.comparables_columns)]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    backend = _select_excel_backend(template_path=template_path)
    if backend == "com":
        metadata_path = _fill_with_excel_com(
            output_path=output_path,
            profile=profile,
            selected=selected,
            target=target,
            visible=visible,
        )
    elif backend == "python-xls":
        metadata_path = _fill_with_python_xls(
            template_path=template_path,
            output_path=output_path,
            profile=profile,
            selected=selected,
            target=target,
        )
        warnings.append(
            "python_xls_backend_flattens_excel_formulas; use Microsoft Excel COM or LibreOffice for formula-safe workbooks"
        )
    elif backend == "libreoffice":
        metadata_path = _fill_with_libreoffice(
            template_path=template_path,
            output_path=output_path,
            profile=profile,
            selected=selected,
            target=target,
        )
    else:
        raise FillError(f"Unsupported Excel backend: {backend}")

    return FillResult(
        output_path=output_path,
        filled_count=len(selected),
        warnings=warnings,
        engine=backend,
        metadata_path=metadata_path,
    )


def _fill_with_excel_com(
    *,
    output_path: Path,
    profile: TemplateProfile,
    selected: list[Comparable],
    target: dict[str, Any] | None,
    visible: bool,
) -> Path | None:
    with ExcelApp(visible=visible) as excel:
        wb = excel.Workbooks.Open(excel_path(output_path), 0, False)
        try:
            ws = wb.Worksheets(profile.sheet_name)
            protected = _protected_cells(profile)
            _fill_comparables(ws, profile, selected, protected)
            if target:
                _fill_target(ws, profile, target, protected)
            wb.Save()
        finally:
            wb.Close(True)
    return None


def _fill_with_libreoffice(
    *,
    template_path: Path,
    output_path: Path,
    profile: TemplateProfile,
    selected: list[Comparable],
    target: dict[str, Any] | None,
) -> Path | None:
    soffice = libreoffice_path()
    if not soffice:
        raise FillError("LibreOffice/soffice is unavailable.")
    if template_path.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
        raise FillError("LibreOffice backend supports .xls, .xlsx, and .xlsm templates.")

    try:
        import openpyxl
    except Exception as exc:
        raise FillError(f"openpyxl is unavailable for LibreOffice backend: {exc}") from exc

    with tempfile.TemporaryDirectory(prefix="realtify_excel_") as tmp_name:
        tmp_dir = Path(tmp_name)
        office_profile = tmp_dir / "lo_profile"
        office_profile.mkdir(parents=True, exist_ok=True)
        source_for_edit = _prepare_xlsx_for_edit(
            soffice=soffice,
            office_profile=office_profile,
            template_path=template_path,
            tmp_dir=tmp_dir,
        )
        edited_xlsx = tmp_dir / "edited.xlsx"
        shutil.copy2(source_for_edit, edited_xlsx)

        workbook = openpyxl.load_workbook(edited_xlsx, keep_vba=template_path.suffix.lower() == ".xlsm", data_only=False)
        try:
            worksheet = _find_openpyxl_sheet(workbook, profile.sheet_name)
            template_rows = _read_template_rows_openpyxl(worksheet, start=15, end=43)
            protected = _protected_cells(profile)
            _fill_comparables_openpyxl(worksheet, profile, selected, protected)
            if target:
                _fill_target_openpyxl(worksheet, profile, target, protected)
            workbook.save(edited_xlsx)
        finally:
            workbook.close()

        if output_path.suffix.lower() in {".xlsx", ".xlsm"}:
            shutil.copy2(edited_xlsx, output_path)
        else:
            converted = _convert_with_libreoffice(
                soffice=soffice,
                office_profile=office_profile,
                input_path=edited_xlsx,
                out_dir=tmp_dir / "final",
                target_format="xls",
            )
            shutil.copy2(converted, output_path)
    payload = build_calculation_sidecar_payload(
        excel_path=output_path,
        engine="libreoffice",
        profile_name=profile.profile,
        candidates=selected,
        target=target,
        template_rows=template_rows,
    )
    return write_excel_sidecar(output_path, payload)


def _fill_with_python_xls(
    *,
    template_path: Path,
    output_path: Path,
    profile: TemplateProfile,
    selected: list[Comparable],
    target: dict[str, Any] | None,
) -> Path:
    if template_path.suffix.lower() != ".xls":
        raise FillError(
            "The cross-platform Python backend currently supports legacy .xls templates only. "
            "Use an .xls template or set REALTIFY_EXCEL_ENGINE=com on Windows."
        )
    try:
        import xlrd
        from xlutils.copy import copy as copy_workbook
    except Exception as exc:
        raise FillError(f"Python XLS backend dependencies are unavailable: {exc}") from exc

    source_book = xlrd.open_workbook(str(template_path), formatting_info=True)
    sheet_index = _find_xlrd_sheet_index(source_book, profile.sheet_name)
    source_sheet = source_book.sheet_by_index(sheet_index)
    template_rows = _read_template_rows(source_sheet, start=15, end=43, date_mode=source_book.datemode)
    output_book = copy_workbook(source_book)
    output_sheet = output_book.get_sheet(sheet_index)
    output_sheet._cell_overwrite_ok = True

    protected = _protected_cells(profile)
    _fill_comparables_xls(output_sheet, profile, selected, protected)
    if target:
        _fill_target_xls(output_sheet, profile, target, protected)

    payload = build_calculation_sidecar_payload(
        excel_path=output_path,
        engine="python-xls",
        profile_name=profile.profile,
        candidates=selected,
        target=target,
        template_rows=template_rows,
    )
    output_book.save(str(output_path))
    return write_excel_sidecar(output_path, payload)


def _validate_candidates(
    candidates: list[Comparable],
    *,
    required_count: int,
    allow_less: bool,
    allow_incomplete: bool,
) -> list[str]:
    warnings: list[str] = []
    if len(candidates) < required_count:
        message = f"only {len(candidates)} candidate(s), required {required_count}"
        if allow_less:
            warnings.append(message)
        else:
            raise FillError(message)

    critical_fields = ["source_url", "address", "area_m2", "price_usd"]
    for idx, candidate in enumerate(candidates[:required_count], start=1):
        missing = [field for field in critical_fields if _candidate_value(candidate, field) in (None, "")]
        if missing:
            message = f"candidate {idx} missing critical field(s): {', '.join(missing)}"
            if allow_incomplete:
                warnings.append(message)
            else:
                raise FillError(message)
    return warnings


def _select_excel_backend(*, template_path: Path) -> str:
    requested = os.getenv("REALTIFY_EXCEL_ENGINE", "auto").strip().lower()
    aliases = {
        "win32": "com",
        "excel": "com",
        "lo": "libreoffice",
        "soffice": "libreoffice",
        "python": "python-xls",
        "xls": "python-xls",
    }
    requested = aliases.get(requested, requested)
    if requested in {"com", "libreoffice", "python-xls"}:
        return requested
    if requested != "auto":
        raise FillError(f"Unsupported REALTIFY_EXCEL_ENGINE value: {requested}")
    if excel_com_available():
        return "com"
    if libreoffice_available():
        return "libreoffice"
    if template_path.suffix.lower() == ".xls":
        return "python-xls"
    raise FillError(
        "No Excel backend is available. Install Microsoft Excel on Windows or LibreOffice/soffice on Linux."
    )


def _fill_comparables(ws: Any, profile: TemplateProfile, candidates: list[Comparable], protected: set[tuple[int, int]]) -> None:
    for col_index, column in enumerate(profile.comparables_columns):
        col_num = _column_to_number(column)
        candidate = candidates[col_index] if col_index < len(candidates) else None
        for field, row in profile.field_mapping.items():
            if (row, col_num) in protected:
                continue
            cell = ws.Cells(row, col_num)
            _clear_cell(cell)
            if not candidate:
                continue
            value = _candidate_value(candidate, field)
            if field == "source_url" and value:
                _write_hyperlink(ws, cell, str(value))
            elif value is not None:
                cell.Value = value


def _fill_comparables_xls(
    ws: Any,
    profile: TemplateProfile,
    candidates: list[Comparable],
    protected: set[tuple[int, int]],
) -> None:
    for col_index, column in enumerate(profile.comparables_columns):
        col_num = _column_to_number(column)
        candidate = candidates[col_index] if col_index < len(candidates) else None
        for field, row in profile.field_mapping.items():
            if (row, col_num) in protected:
                continue
            value = _candidate_value(candidate, field) if candidate else None
            _write_xls_cell(ws, row, col_num, value)


def _fill_target(ws: Any, profile: TemplateProfile, target: dict[str, Any], protected: set[tuple[int, int]]) -> None:
    if not profile.target_column:
        return
    col_num = _column_to_number(profile.target_column)
    for field, row in profile.field_mapping.items():
        if field == "source_url" or (row, col_num) in protected:
            continue
        cell = ws.Cells(row, col_num)
        _clear_cell(cell)
        value = _target_value(target, field)
        if value is not None:
            cell.Value = value


def _fill_target_xls(ws: Any, profile: TemplateProfile, target: dict[str, Any], protected: set[tuple[int, int]]) -> None:
    if not profile.target_column:
        return
    col_num = _column_to_number(profile.target_column)
    for field, row in profile.field_mapping.items():
        if field == "source_url" or (row, col_num) in protected:
            continue
        _write_xls_cell(ws, row, col_num, _target_value(target, field))


def _fill_comparables_openpyxl(
    ws: Any,
    profile: TemplateProfile,
    candidates: list[Comparable],
    protected: set[tuple[int, int]],
) -> None:
    for col_index, column in enumerate(profile.comparables_columns):
        col_num = _column_to_number(column)
        candidate = candidates[col_index] if col_index < len(candidates) else None
        for field, row in profile.field_mapping.items():
            if (row, col_num) in protected:
                continue
            value = _candidate_value(candidate, field) if candidate else None
            _write_openpyxl_cell(ws.cell(row=row, column=col_num), value, hyperlink=field == "source_url")


def _fill_target_openpyxl(ws: Any, profile: TemplateProfile, target: dict[str, Any], protected: set[tuple[int, int]]) -> None:
    if not profile.target_column:
        return
    col_num = _column_to_number(profile.target_column)
    for field, row in profile.field_mapping.items():
        if field == "source_url" or (row, col_num) in protected:
            continue
        _write_openpyxl_cell(ws.cell(row=row, column=col_num), _target_value(target, field))


def _write_openpyxl_cell(cell: Any, value: Any, *, hyperlink: bool = False) -> None:
    cell.value = "" if value is None else value
    if hyperlink and value:
        cell.hyperlink = str(value)
        cell.style = "Hyperlink"
    elif hyperlink:
        cell.hyperlink = None


def _write_xls_cell(ws: Any, row: int, col: int, value: Any) -> None:
    if value is None:
        value = ""
    if hasattr(value, "unicode_string"):
        value = str(value)
    ws.write(row - 1, col - 1, value)


def _prepare_xlsx_for_edit(
    *,
    soffice: str,
    office_profile: Path,
    template_path: Path,
    tmp_dir: Path,
) -> Path:
    suffix = template_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        target = tmp_dir / template_path.name
        shutil.copy2(template_path, target)
        return target
    return _convert_with_libreoffice(
        soffice=soffice,
        office_profile=office_profile,
        input_path=template_path,
        out_dir=tmp_dir / "xlsx",
        target_format="xlsx",
    )


def _convert_with_libreoffice(
    *,
    soffice: str,
    office_profile: Path,
    input_path: Path,
    out_dir: Path,
    target_format: str,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    command = [
        soffice,
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--nolockcheck",
        f"-env:UserInstallation={office_profile.as_uri()}",
        "--convert-to",
        target_format,
        "--outdir",
        str(out_dir),
        str(input_path.resolve()),
    ]
    result = subprocess.run(command, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        details = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part)
        raise FillError(f"LibreOffice conversion to {target_format} failed: {details}")
    suffix = "." + target_format.split(":", 1)[0].lower()
    matches = sorted(out_dir.glob(f"*{suffix}"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not matches:
        details = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part)
        raise FillError(f"LibreOffice did not create a {suffix} file. {details}")
    return matches[0]


def _find_openpyxl_sheet(workbook: Any, preferred_name: str) -> Any:
    if preferred_name in workbook.sheetnames:
        return workbook[preferred_name]
    preferred = preferred_name.casefold()
    for name in workbook.sheetnames:
        if preferred and preferred in name.casefold():
            return workbook[name]
    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 44 and worksheet.max_column >= 8:
            return worksheet
    return workbook.worksheets[0]


def _find_xlrd_sheet_index(book: Any, preferred_name: str) -> int:
    names = book.sheet_names()
    for index, name in enumerate(names):
        if name == preferred_name:
            return index
    preferred = preferred_name.casefold()
    for index, name in enumerate(names):
        if preferred and preferred in name.casefold():
            return index
    for index, sheet in enumerate(book.sheets()):
        if sheet.nrows >= 44 and sheet.ncols >= 8:
            return index
    return 0


def _read_template_rows(sheet: Any, *, start: int, end: int, date_mode: int) -> dict[int, list[Any]]:
    rows: dict[int, list[Any]] = {}
    for row_index in range(start, end + 1):
        row: list[Any] = []
        for col_index in range(1, 9):
            row.append(_xlrd_cell_value(sheet, row_index, col_index, date_mode=date_mode))
        rows[row_index] = row
    return rows


def _read_template_rows_openpyxl(sheet: Any, *, start: int, end: int) -> dict[int, list[Any]]:
    rows: dict[int, list[Any]] = {}
    for row_index in range(start, end + 1):
        row: list[Any] = []
        for col_index in range(1, 9):
            row.append(sheet.cell(row=row_index, column=col_index).value or "")
        rows[row_index] = row
    return rows


def _xlrd_cell_value(sheet: Any, row: int, col: int, *, date_mode: int) -> Any:
    try:
        cell = sheet.cell(row - 1, col - 1)
    except IndexError:
        return ""
    if cell.ctype == 0:
        return ""
    if cell.ctype == 3:
        try:
            import xlrd

            return xlrd.xldate.xldate_as_datetime(cell.value, date_mode).date()
        except Exception:
            return cell.value
    return cell.value


def _candidate_value(candidate: Comparable, field: str) -> Any:
    if field == "address":
        address = candidate.address
        if address and candidate.complex_name and candidate.complex_name not in address:
            return f"{address}, {candidate.complex_name}"
        return address
    if field == "price_usd":
        if candidate.price_usd is not None:
            return candidate.price_usd
        if candidate.currency == "USD" and candidate.price is not None:
            return candidate.price
        return None
    if field == "purpose":
        if candidate.purpose:
            return candidate.purpose
        if candidate.property_type == "parking":
            return "паркінг"
        return None
    return getattr(candidate, field, None)


def _target_value(target: dict[str, Any], field: str) -> Any:
    if field == "address":
        address = target.get("address")
        complex_name = target.get("complex_name")
        if address and complex_name and complex_name not in address:
            return f"{address}, {complex_name}"
        return address
    if field == "purpose" and not target.get(field):
        if target.get("property_type") == "parking":
            return "паркінг"
    return target.get(field)


def _write_hyperlink(ws: Any, cell: Any, url: str) -> None:
    cell.Value = url
    try:
        cell.Hyperlinks.Delete()
    except Exception:
        pass
    ws.Hyperlinks.Add(Anchor=cell, Address=url, TextToDisplay=url)


def _clear_cell(cell: Any) -> None:
    try:
        cell.Hyperlinks.Delete()
    except Exception:
        pass
    cell.ClearContents()


def _protected_cells(profile: TemplateProfile) -> set[tuple[int, int]]:
    protected: set[tuple[int, int]] = set()
    for range_ref in profile.protected_cells:
        protected.update(_expand_a1_range(range_ref))
    return protected


def _expand_a1_range(range_ref: str) -> set[tuple[int, int]]:
    parts = range_ref.split(":")
    start = _parse_cell_ref(parts[0])
    end = _parse_cell_ref(parts[-1])
    cells: set[tuple[int, int]] = set()
    for row in range(min(start[0], end[0]), max(start[0], end[0]) + 1):
        for col in range(min(start[1], end[1]), max(start[1], end[1]) + 1):
            cells.add((row, col))
    return cells


def _parse_cell_ref(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", value.strip().upper())
    if not match:
        raise FillError(f"Invalid cell reference: {value}")
    return int(match.group(2)), _column_to_number(match.group(1))


def _column_to_number(column: str) -> int:
    number = 0
    for char in column.strip().upper():
        if not ("A" <= char <= "Z"):
            raise FillError(f"Invalid column: {column}")
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _load_task(path: Path | None) -> dict[str, Any]:
    if not path:
        return {}
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _default_output_path(profile_name: str) -> Path:
    output_dir = ensure_output_dir(datetime.now().strftime("%Y%m%d_%H%M%S_excel"))
    return output_dir / f"{profile_name}_filled.xls"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Fill Excel valuation template from candidates.json.")
    parser.add_argument("--task", type=Path, default=None, help="Optional task YAML with template and target sections.")
    parser.add_argument("--candidates", type=Path, required=True, help="Path to candidates.json from collect_from_links.")
    parser.add_argument("--template", type=Path, default=None, help="Path to .xls template. Overrides task.template.path.")
    parser.add_argument("--profile", default=None, help="Template profile name, e.g. parking or apartment.")
    parser.add_argument("--profile-path", type=Path, default=None, help="Path to template profile YAML.")
    parser.add_argument("--out", type=Path, default=None, help="Output .xls path.")
    parser.add_argument("--required-count", type=int, default=None)
    parser.add_argument("--allow-less", action="store_true", help="Allow fewer candidates than required.")
    parser.add_argument("--allow-incomplete", action="store_true", help="Allow missing critical fields.")
    parser.add_argument("--visible", action="store_true", help="Show Excel while filling.")
    args = parser.parse_args(argv)

    console = Console()
    try:
        task = _load_task(args.task)
        task_template = (task.get("template") or {}) if isinstance(task.get("template"), dict) else {}
        profile_name = args.profile or task_template.get("profile")
        if not profile_name and not args.profile_path:
            raise FillError("Template profile is required. Use --profile or --profile-path.")

        profile_path = args.profile_path or RESOURCE_ROOT / "config" / "template_profiles" / f"{profile_name}.yaml"
        profile = load_template_profile(profile_path)
        template_path = args.template or (Path(task_template["path"]) if task_template.get("path") else None)
        if not template_path:
            raise FillError("Template path is required. Use --template or --task with template.path.")

        candidates = load_candidates(args.candidates)
        output_path = args.out or _default_output_path(profile.profile)
        target = task.get("target") if isinstance(task.get("target"), dict) else None
        result = fill_excel_template(
            template_path=template_path,
            profile=profile,
            candidates=candidates,
            output_path=output_path,
            target=target,
            required_count=args.required_count,
            allow_less=args.allow_less,
            allow_incomplete=args.allow_incomplete,
            visible=args.visible,
        )
    except Exception as exc:
        console.print(f"[red]Excel fill failed:[/red] {exc}")
        return 1

    console.print(f"[green]Filled {result.filled_count} comparable(s)[/green]")
    if result.warnings:
        for warning in result.warnings:
            console.print(f"[yellow]Warning:[/yellow] {warning}")
    console.print(f"Output: {result.output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
