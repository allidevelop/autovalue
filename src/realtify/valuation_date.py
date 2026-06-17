from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from realtify import nbu_rate as nbu_rate_api
from realtify import valuation_register
from realtify.paths import PROJECT_ROOT
from realtify.valuation_register import RegisterEntry

# Орієнтовний резервний курс — лише якщо реєстр/мережа недоступні.
DEFAULT_FALLBACK_NBU_RATE = 41.0


@dataclass
class ValuationContext:
    """Дата оцінки + курс НБУ на неї + примітки для лога/звіту."""

    valuation_date: date
    nbu_rate: float | None
    rate_is_official: bool
    source: str  # 'explicit' | 'register' | 'cell' | 'today'
    matched_entry: RegisterEntry | None = None
    notes: list[str] = field(default_factory=list)


def resolve_valuation_date(*, task: dict[str, Any], excel_path: Path | None = None) -> date:
    target = task.get("target") if isinstance(task.get("target"), dict) else {}
    explicit = _parse_date(_first_not_empty(task.get("valuation_date"), target.get("valuation_date")))
    if explicit:
        return explicit

    from_register, _entry = _resolve_from_register(task)
    if from_register:
        return from_register

    source = _first_mapping(
        task.get("valuation_date_source"),
        target.get("valuation_date_source"),
        (task.get("template") or {}).get("valuation_date_source") if isinstance(task.get("template"), dict) else None,
    )
    if source:
        template = task.get("template") if isinstance(task.get("template"), dict) else {}
        if source.get("path") in (None, "", "template", "uploaded_excel") and template.get("path"):
            source = {**source, "path": template.get("path")}
        source_value = _read_valuation_date_source(source, fallback_excel_path=excel_path)
        parsed = _parse_date(source_value)
        if parsed:
            return parsed

    return date.today()


def resolve_valuation_context(task: dict[str, Any], *, excel_path: Path | None = None) -> ValuationContext:
    """Повний контекст оцінки: дата + курс НБУ на цю дату + примітки.

    Порядок дати: явна → реєстр клієнта → excel-комірка → сьогодні.
    Курс: офіційний НБУ на дату; якщо мережа/реєстр недоступні — орієнтовний резерв.
    """
    target = task.get("target") if isinstance(task.get("target"), dict) else {}
    notes: list[str] = []

    explicit = _parse_date(_first_not_empty(task.get("valuation_date"), target.get("valuation_date")))
    if explicit:
        valuation_date, source, entry = explicit, "explicit", None
    else:
        register_date, entry = _resolve_from_register(task)
        if register_date and entry is not None:
            valuation_date, source = register_date, "register"
            notes.append(
                f"Дата оцінки з реєстру: {register_date.strftime('%d.%m.%Y')} "
                f"(кв. {entry.apartment or '?'}, {entry.address or '?'}"
                + (f", {entry.fund}" if entry.fund else "")
                + ")."
            )
        else:
            cell_date = resolve_valuation_date(task=task, excel_path=excel_path)
            if cell_date == date.today():
                valuation_date, source = cell_date, "today"
                if _register_configured(task):
                    notes.append(
                        "Реєстр оцінок: об'єкт не знайдено за № квартири/адресою — "
                        f"дату взято поточну ({cell_date.strftime('%d.%m.%Y')})."
                    )
            else:
                valuation_date, source = cell_date, "cell"
            entry = None

    rate = nbu_rate_api.usd_uah_rate(valuation_date)
    if rate is not None:
        rate_is_official = True
        notes.append(f"Курс НБУ на {valuation_date.strftime('%d.%m.%Y')}: {rate:.4f} грн/USD.")
    else:
        rate = DEFAULT_FALLBACK_NBU_RATE
        rate_is_official = False
        notes.append(
            f"Курс НБУ недоступний — використано орієнтовний резерв {rate:.2f} грн/USD."
        )

    return ValuationContext(
        valuation_date=valuation_date,
        nbu_rate=rate,
        rate_is_official=rate_is_official,
        source=source,
        matched_entry=entry,
        notes=notes,
    )


def _resolve_from_register(task: dict[str, Any]) -> tuple[date | None, RegisterEntry | None]:
    path = valuation_register.register_path_from_task(task)
    if not path:
        return None, None
    target = task.get("target") if isinstance(task.get("target"), dict) else {}
    entries = valuation_register.load_register(path)
    if not entries:
        return None, None
    entry = valuation_register.find_entry(
        entries,
        apartment=target.get("apartment_number") or target.get("apartment"),
        city=target.get("city"),
        address=target.get("address"),
    )
    if entry and entry.valuation_date:
        return entry.valuation_date, entry
    return None, None


def _register_configured(task: dict[str, Any]) -> bool:
    return valuation_register.register_path_from_task(task) is not None


def _read_valuation_date_source(source: dict[str, Any], *, fallback_excel_path: Path | None) -> Any:
    source_type = str(source.get("type") or "excel_cell")
    if source_type != "excel_cell":
        return None

    path_value = source.get("path")
    if path_value in (None, "", "excel", "template", "uploaded_excel"):
        path = fallback_excel_path
    else:
        path = Path(str(path_value))
    if path is None:
        return None
    path = path if path.is_absolute() else PROJECT_ROOT / path
    if not path.exists():
        return None

    cell_ref = str(source.get("cell") or "").strip()
    if not cell_ref:
        return None
    sheet = source.get("sheet")

    try:
        if path.suffix.lower() == ".xls":
            return _read_xls_cell(path, sheet_name=str(sheet) if sheet else None, cell_ref=cell_ref)
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return _read_xlsx_cell(path, sheet_name=str(sheet) if sheet else None, cell_ref=cell_ref)
    except Exception:
        return None
    return None


def _read_xls_cell(path: Path, *, sheet_name: str | None, cell_ref: str) -> Any:
    try:
        import xlrd
    except Exception:
        return None
    book = xlrd.open_workbook(str(path))
    sheet = _xlrd_sheet(book, sheet_name)
    row, col = _parse_cell_ref(cell_ref)
    cell = sheet.cell(row - 1, col - 1)
    if cell.ctype == 3:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode).date()
        except Exception:
            return cell.value
    return cell.value


def _read_xlsx_cell(path: Path, *, sheet_name: str | None, cell_ref: str) -> Any:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        return sheet[cell_ref].value
    finally:
        workbook.close()


def _xlrd_sheet(book: Any, sheet_name: str | None) -> Any:
    if sheet_name:
        for name in book.sheet_names():
            if name == sheet_name:
                return book.sheet_by_name(name)
    return book.sheet_by_index(0)


def _parse_cell_ref(value: str) -> tuple[int, int]:
    import re

    match = re.fullmatch(r"([A-Za-z]+)(\d+)", value.strip())
    if not match:
        raise ValueError(f"Invalid Excel cell reference: {value}")
    column = 0
    for char in match.group(1).upper():
        column = column * 26 + (ord(char) - ord("A") + 1)
    return int(match.group(2)), column


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d,%m,%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _first_not_empty(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _first_mapping(*values: Any) -> dict[str, Any] | None:
    for value in values:
        if isinstance(value, dict):
            return value
    return None
