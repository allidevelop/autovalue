"""Реєстр оцінок клієнта (Excel «Продаж квартир»): джерело дати оцінки.

Кожен рядок = об'єкт на оцінку: № квартири, дата оцінки, місто, адреса, фонд (ЖК).
Матчинг об'єкта (з PDF-intake) до рядка реєстру — детермінований, за нормалізованою
адресою будинку (`analog_cache.address_key`) + номером квартири. Звідси беремо
дату оцінки, а далі — курс НБУ на цю дату (`nbu_rate`).
"""

from __future__ import annotations

import os
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from realtify.analog_cache import address_key, building_match_key
from realtify.paths import PROJECT_ROOT

ENV_REGISTER_PATH = "REALTIFY_VALUATION_REGISTER"

# Куди зберігати/звідки читати керований реєстр, якщо env не задано (UI-завантаження).
DEFAULT_MANAGED_REGISTER = PROJECT_ROOT / "data" / "registers" / "valuation_register.xlsx"

# Куди писати результат розрахунку — окремі колонки, без зачіпання клієнтської «Ціна продажу».
RESULT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Оцінка системи, грн", "estimate"),
    ("Курс НБУ", "rate"),
    ("Дата розрахунку", "calc_date"),
)

# Гнучке зіставлення заголовків колонок (укр/рос/en).
_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "apartment": ("номер квартири", "№ квартири", "номер кв", "квартира", "апартамент", "apartment", "номер"),
    "date": ("дата оцінки", "дата оценки", "дата", "date"),
    "price": ("ціна продажу", "цена продажи", "ціна", "цена", "price"),
    "city": ("місто", "город", "city"),
    "address": ("адреса", "адрес", "address"),
    "fund": ("фонд", "жк", "комплекс", "fund", "complex"),
}


@dataclass(frozen=True)
class RegisterEntry:
    apartment: str | None
    city: str | None
    address: str | None
    fund: str | None
    valuation_date: date | None
    price_uah: float | None
    sheet: str
    row: int


def register_path_from_task(task: dict[str, Any] | None) -> Path | None:
    """Шлях до реєстру: спершу task.valuation.register_path, далі — env."""
    if isinstance(task, dict):
        valuation = task.get("valuation")
        if isinstance(valuation, dict):
            raw = valuation.get("register_path") or valuation.get("register")
            resolved = _resolve_path(raw)
            if resolved:
                return resolved
    env = register_path_from_env()
    if env:
        return env
    # Керований реєстр, завантажений через UI (якщо env не задано).
    return DEFAULT_MANAGED_REGISTER if DEFAULT_MANAGED_REGISTER.exists() else None


def register_path_from_env() -> Path | None:
    return _resolve_path(os.environ.get(ENV_REGISTER_PATH))


def register_path_default() -> Path:
    """Канонічний шлях керованого реєстру (куди пише UI-завантажувач і звідки
    читає резолвер): env-шлях, якщо задано, інакше дефолт у data/registers."""
    return register_path_from_env() or DEFAULT_MANAGED_REGISTER


def load_register(path: Path) -> list[RegisterEntry]:
    """Зчитує всі аркуші реєстру у плоский список об'єктів."""
    try:
        from openpyxl import load_workbook
    except Exception:  # noqa: BLE001 — без openpyxl реєстр недоступний
        return []
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    entries: list[RegisterEntry] = []
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            header_index, column_map = _detect_header(rows)
            if column_map is None:
                continue
            for offset, row in enumerate(rows[header_index + 1:], start=header_index + 2):
                entry = _row_to_entry(row, column_map, sheet=worksheet.title, row_number=offset)
                if entry is not None:
                    entries.append(entry)
    finally:
        workbook.close()
    return entries


def find_entry(
    entries: list[RegisterEntry],
    *,
    apartment: str | None,
    city: str | None,
    address: str | None,
) -> RegisterEntry | None:
    """Знаходить рядок реєстру для об'єкта. Неоднозначність → None (чесно)."""
    want_addr = _building_key(city, address)
    if not want_addr:
        return None
    want_apt = _norm_apartment(apartment)

    same_building = [e for e in entries if _building_key(e.city, e.address) == want_addr]
    if not same_building:
        return None

    if want_apt:
        exact = [e for e in same_building if _norm_apartment(e.apartment) == want_apt]
        if len(exact) == 1:
            return exact[0]
        if len(exact) > 1:
            # Кілька записів на ту саму квартиру — беремо найсвіжіший за датою.
            return max(exact, key=lambda e: e.valuation_date or date.min)
        return None  # квартира відома, але в реєстрі її немає → не вгадуємо

    # Номер квартири невідомий: однозначно лише якщо у будинку рівно один запис.
    return same_building[0] if len(same_building) == 1 else None


def write_estimate(
    path: Path,
    entry: RegisterEntry,
    *,
    estimate_uah: float | None,
    nbu_rate: float | None,
    calc_date: date,
    backup: bool = True,
) -> bool:
    """Записує результат оцінки в окремі колонки реєстру (рядок entry.row).

    НЕ чіпає клієнтську «Ціна продажу». Перед першим записом робить .backup-копію.
    Повертає True, якщо записано й збережено.
    """
    try:
        from openpyxl import load_workbook
    except Exception:  # noqa: BLE001 — без openpyxl запис недоступний
        return False
    if not path.exists():
        return False
    # openpyxl не round-trip'ить макроси/бінарний формат — не чіпаємо такі книги.
    if path.suffix.lower() in {".xlsm", ".xlsb", ".xls"}:
        return False

    if backup:
        backup_path = path.with_name(f"{path.stem}.backup{path.suffix}")
        if not backup_path.exists():
            try:
                shutil.copy2(path, backup_path)
            except Exception:  # noqa: BLE001 — бекап не критичний для запису
                pass

    workbook = load_workbook(path)  # не read_only — зберігаємо форматування
    try:
        # entry.row — абсолютний індекс рядка В СВОЄМУ аркуші. Якщо аркуш зник —
        # НЕ пишемо в чужий лист (інакше затремо неповʼязані дані), а відмовляємось.
        if entry.sheet not in workbook.sheetnames:
            return False
        sheet = workbook[entry.sheet]
        header_row = _find_header_row_ws(sheet)
        if header_row is None:
            return False

        existing: dict[str, int] = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(header_row, col).value
            if value not in (None, ""):
                existing[str(value).strip()] = col

        col_for: dict[str, int] = {}
        next_col = sheet.max_column + 1
        for header, key in RESULT_COLUMNS:
            if header in existing:
                col_for[key] = existing[header]
            else:
                sheet.cell(header_row, next_col, header)
                col_for[key] = next_col
                next_col += 1

        if estimate_uah is not None:
            sheet.cell(entry.row, col_for["estimate"], round(float(estimate_uah), 2))
        if nbu_rate is not None:
            sheet.cell(entry.row, col_for["rate"], round(float(nbu_rate), 4))
        sheet.cell(entry.row, col_for["calc_date"], calc_date)

        workbook.save(path)
        return True
    finally:
        workbook.close()


# ── допоміжне ──────────────────────────────────────────────────────────────


def _find_header_row_ws(sheet: Any) -> int | None:
    max_col = min(sheet.max_column, 64)
    for row_number in range(1, 7):
        row = tuple(sheet.cell(row_number, col).value for col in range(1, max_col + 1))
        mapping = _match_header_row(row)
        if mapping and "date" in mapping and "address" in mapping:
            return row_number
    return None

def _resolve_path(raw: Any) -> Path | None:
    if not raw:
        return None
    path = Path(str(raw))
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path if path.exists() else None


def _building_key(city: str | None, address: str | None) -> str:
    if not address:
        return ""
    # Толерантний матч-ключ (знімає область/тип вулиці/дублі міста) — інакше
    # «Львівська обл., м. Львів, вулиця Хмельницького…» не збігалося з реєстром.
    return building_match_key(city=city, address=address)


def normalize_apartment(value: Any) -> str:
    """Публічна нормалізація № квартири (для звірки точного збігу ззовні)."""
    return _norm_apartment(value)


def _norm_apartment(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip().lower()
    if text.endswith(".0"):  # 353.0 з float-комірки
        text = text[:-2]
    return re.sub(r"[^0-9a-zа-яіїєґ]", "", text)


def _detect_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int] | None]:
    for index, row in enumerate(rows[:6]):
        mapping = _match_header_row(row)
        if mapping and "date" in mapping and "address" in mapping:
            return index, mapping
    return 0, None


def _match_header_row(row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_index, cell in enumerate(row):
        if cell in (None, ""):
            continue
        text = str(cell).strip().lower()
        for field, aliases in _HEADER_ALIASES.items():
            if field in mapping:
                continue
            if text in aliases or any(text.startswith(alias) for alias in aliases):
                mapping[field] = col_index
                break
    return mapping


def _row_to_entry(
    row: tuple[Any, ...],
    column_map: dict[str, int],
    *,
    sheet: str,
    row_number: int,
) -> RegisterEntry | None:
    def cell(field: str) -> Any:
        col = column_map.get(field)
        if col is None or col >= len(row):
            return None
        return row[col]

    valuation_date = _parse_date(cell("date"))
    apartment = _stringify(cell("apartment"))
    address = _stringify(cell("address"))
    if valuation_date is None and apartment is None and address is None:
        return None  # порожній рядок
    return RegisterEntry(
        apartment=apartment,
        city=_stringify(cell("city")),
        address=address,
        fund=_stringify(cell("fund")),
        valuation_date=valuation_date,
        price_uah=_parse_money(cell("price")),
        sheet=sheet,
        row=row_number,
    )


def _stringify(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text or None


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d,%m,%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None
