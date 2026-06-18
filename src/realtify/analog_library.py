"""Масовий імпорт бібліотеки аналогів (адреса → посилання на аналоги).

Клієнт/оцінювач дає файл (CSV/Excel/текст) зі стовпцями `address` та `url`
(плюс опційно `city`, `property_type`, `complex_name`). Для кожної адреси система
один раз збирає аналоги (дані + скриншоти) і кладе їх у бібліотеку (analog_cache),
яка далі перевикористовується в оцінках цього будинку — без повторного пошуку.

Формат файлу (заголовки гнучкі, укр/рус/eng):
    address ; url ; city ; property_type ; complex_name
Один рядок = один аналог; кілька рядків з тією ж адресою = аналоги цього будинку.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from realtify import analog_cache
from realtify import report_db
from realtify.collect_from_links import collect_from_links
from realtify.models import PropertyType, TransactionType
from realtify.progress import ProgressCallback, emit_progress
from realtify.source_config import load_sources_config

_HEADER_MAP = {
    "address": "address", "адреса": "address", "адрес": "address",
    "обʼєкт": "address", "объект": "address", "будинок": "address",
    "url": "url", "посилання": "url", "ссылка": "url", "link": "url", "аналог": "url",
    "city": "city", "місто": "city", "город": "city",
    "property_type": "property_type", "тип": "property_type", "тип обʼєкта": "property_type",
    "complex_name": "complex_name", "жк": "complex_name", "комплекс": "complex_name",
}


@dataclass
class LibraryEntry:
    address: str
    urls: list[str] = field(default_factory=list)
    city: str | None = None
    property_type: str = "apartment"
    transaction_type: str = "sale"
    complex_name: str | None = None


def _norm_header(value: Any) -> str:
    key = str(value or "").strip().casefold()
    return _HEADER_MAP.get(key, key)


def _read_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    lines = text.splitlines()
    if not lines:
        return []
    first = lines[0]
    delimiter = ";" if ";" in first else ("\t" if "\t" in first else ",")
    reader = csv.DictReader(lines, delimiter=delimiter)
    rows: list[dict[str, str]] = []
    for raw in reader:
        rows.append({_norm_header(k): str(v or "").strip() for k, v in raw.items() if k})
    return rows


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Для .xlsx потрібен openpyxl; або експортуйте файл у CSV.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    grid = list(ws.iter_rows(values_only=True))
    if not grid:
        return []
    headers = [_norm_header(c) for c in grid[0]]
    rows: list[dict[str, str]] = []
    for line in grid[1:]:
        row = {
            headers[i]: (str(line[i]).strip() if i < len(line) and line[i] is not None else "")
            for i in range(len(headers))
        }
        rows.append(row)
    return rows


def parse_library_file(path: Path) -> list[LibraryEntry]:
    """Файл → перелік адрес з їх аналог-посиланнями (згруповано за адресою)."""
    grouped: dict[str, LibraryEntry] = {}
    for row in _read_rows(path):
        address = (row.get("address") or "").strip()
        url = (row.get("url") or "").strip()
        if not address or not url.startswith("http"):
            continue
        key = address.casefold()
        entry = grouped.get(key)
        if entry is None:
            entry = LibraryEntry(
                address=address,
                city=(row.get("city") or "").strip() or None,
                property_type=(row.get("property_type") or "apartment").strip() or "apartment",
                complex_name=(row.get("complex_name") or "").strip() or None,
            )
            grouped[key] = entry
        if url not in entry.urls:
            entry.urls.append(url)
    return list(grouped.values())


def import_library(
    entries: list[LibraryEntry],
    *,
    output_dir: Path,
    progress: ProgressCallback | None = None,
) -> dict[str, Any]:
    """Збирає аналоги по кожній адресі та зберігає у бібліотеку (analog_cache)."""
    sources = load_sources_config(None)
    report: dict[str, Any] = {
        "addresses": len(entries),
        "saved_addresses": 0,
        "saved_analogs": 0,
        "results": [],
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    for index, entry in enumerate(entries, start=1):
        prop = entry.property_type if entry.property_type in PropertyType.__args__ else "apartment"
        trans = entry.transaction_type if entry.transaction_type in TransactionType.__args__ else "sale"
        emit_progress(
            progress,
            f"[{index}/{len(entries)}] {entry.address}: збираю {len(entry.urls)} аналогів…",
        )
        item: dict[str, Any] = {"address": entry.address, "urls": len(entry.urls)}
        try:
            sub = output_dir / f"addr_{index:03d}"
            sub.mkdir(parents=True, exist_ok=True)
            collection = collect_from_links(
                entry.urls,
                output_dir=sub,
                sources_config=sources,
                property_type=prop,
                transaction_type=trans,
                progress=progress,
            )
            if not collection.candidates:
                item["status"] = "no_candidates"
                item["collected"] = 0
                report["results"].append(item)
                emit_progress(progress, f"[{index}/{len(entries)}] {entry.address}: 0 зібрано — пропускаю.")
                continue
            # Пишемо в НОВУ курировану базу report_comparables (та сама, що CRUD/
            # add-by-URL/імпорт звітів), а не в старий analog_cache. Аналог — за
            # власною (зібраною) адресою; адреса/місто/ЖК із файлу — як фолбек.
            rows: list[dict[str, Any]] = []
            for cand in collection.candidates:
                addr = cand.address or entry.address
                ak = analog_cache.address_key(
                    city=cand.city or entry.city, address=addr, property_type=prop, complex_name=None,
                )
                dk = report_db.compute_dedup_key(ak, cand.area_m2, cand.price_usd, cand.floor_or_level)
                sp = None
                img = cand.report_image_path or cand.screenshot_path
                if img and Path(str(img)).exists():
                    try:
                        sp = str(report_db.store_screenshot(dk, Path(str(img)).read_bytes()))
                    except Exception:  # noqa: BLE001
                        sp = None
                rows.append({
                    "city": cand.city or entry.city,
                    "address": addr,
                    "complex_name": cand.complex_name or entry.complex_name,
                    "property_type": prop,
                    "area_m2": cand.area_m2,
                    "price_usd": cand.price_usd,
                    "price_per_m2_usd": cand.price_per_m2_usd,
                    "floor_or_level": cand.floor_or_level,
                    "rooms": cand.rooms,
                    "location_quality": cand.location_quality,
                    "building_class": cand.building_class,
                    "condition": cand.condition,
                    "delivery_date": cand.delivery_date,
                    "source_key": "library",
                    "source_url": str(cand.source_url) if cand.source_url else None,
                    "screenshot_path": sp,
                })
            stats = report_db.upsert_many(rows)
            item["status"] = "saved"
            item["collected"] = len(rows)
            item["db"] = stats
            report["saved_addresses"] += 1
            report["saved_analogs"] += len(rows)
            emit_progress(
                progress,
                f"[{index}/{len(entries)}] {entry.address}: у базу {len(rows)} аналогів "
                f"(нових {stats.get('inserted', 0)}, оновлено {stats.get('updated', 0)}).",
            )
        except Exception as exc:  # noqa: BLE001 — одна адреса не повинна валити весь імпорт
            item["status"] = "error"
            item["error"] = str(exc)
            emit_progress(progress, f"[{index}/{len(entries)}] {entry.address}: помилка — {exc}")
        report["results"].append(item)
    return report
